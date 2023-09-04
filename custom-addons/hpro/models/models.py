# -*- coding: utf-8 -*-
import base64
import io
import openpyxl
from collections import defaultdict

from odoo.exceptions import ValidationError
from odoo import models, fields, api





class on_work(models.Model):
    _name = 'on.work'
    _description = '现场工序'
    _rec_name = 'employee'
    _order = "date1 desc"

    date1 = fields.Date(string='日期', required=True)
    week = fields.Char(string="周", compute="set_week", store=True)
    @api.depends("date1")
    def set_week(self):
        for record in self:
            if record.date1:
                record.week = f"{record.date1.year}年第{record.date1.isocalendar()[1]}周"


    order_no = fields.Many2one('sale_pro.sale_pro', string='订单号', required=True)
    order_number = fields.Many2one('ib.detail', string='款式编号', required=True)

    employee_id = fields.Integer(string='工序号', required=True)
    part_name = fields.Char(string='部件名称')
    process_abbreviation = fields.Char(string='工序描述')
    mechanical_type = fields.Char(string='机器类型')
    process_level = fields.Char(string='工序等级')
    standard_time = fields.Float(string='标准时间')
    standard_price = fields.Float(string='原单价', digits=(16, 3), compute="set_standard_price", store=True)

    # @api.depends("order_number", "employee_id")
    # def set_standard_price(self):
    #     for record in self:
    #         work_work_obj = self.env['work.work'].sudo().search([("order_number", "=", record.order_number.id), ("employee_id", "=", record.employee_id)])
    #         if work_work_obj:
    #             record.standard_price = work_work_obj.standard_price


            # else:
            #     raise ValidationError(f"工序单中没有查询到款号{record.order_number.style_number},工序号{record.employee_id}")


    label = fields.Text(string='标签')
    employee = fields.Many2one('hr.employee', string='员工', required=True)
    employee_id2 = fields.Char(string='员工id')
    group = fields.Char(string='组别', required=True)
    over_number = fields.Float('件数', required=True)
    state = fields.Selection([
        ('已确认', '已确认'),
        ('未确认', '未确认')
        ], string="状态", default="未确认")
    made_employee_id = fields.Many2one("memp.memp", string="员工信息")
    piecerate_id = fields.Many2one("cost.cost1", string="工序工资")
    ji_jian_id = fields.Many2one("ji.jian", string="计件工资")
    ji_jian_week_id = fields.Many2one("ji.jian.week", string="计件工资(周)")
    eff_eff_id = fields.Many2one("eff.eff", string="个人效率表(日)")
    group_efficiency_id = fields.Many2one("group.efficiency", string="组效率表(日)")
    eff_eff_week_id = fields.Many2one("eff.eff.week", string="个人效率表(周)")
    group_efficiency_week_id = fields.Many2one("group.efficiency.week", string="组效率表(周)")
    contract_type = fields.Selection(string="合同", related="employee.is_it_a_temporary_worker", store=True)

    # 辅助字段
    non_conformance = fields.Boolean(string='数量', default=False)
    non_price = fields.Boolean(string='价格', default=False)

    @api.model
    def process_formal_excel_data(self, data):
        """正式工导入读取excel文件生成"""
        decoded_data = base64.b64decode(data.split(',')[1])
        file_object = io.BytesIO(decoded_data)
        wb = openpyxl.load_workbook(file_object)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True, min_row=2):
            # 获取列
            process_number = row[0]
            style_number = row[2]

            # 如果工序号和款式标号都存在且不为空，则存入相应的列表
            if process_number and style_number:
                work = self.env['work.work'].search([('employee_id', '=', process_number), ('order_number', '=', style_number)])
                if not work:
                    raise ValidationError(f"在工序单中没有搜索到Excel表中对应的工序号：{process_number}, 款号:{style_number} ！")
                else:
                    for matched_record in work:
                        # 查询订单号
                        order_number_id = self.env['sale_pro.sale_pro'].search([('order_number', '=', row[1])]).id
                        # 查询款号
                        style_no_id = self.env['ib.detail'].search([('style_number', '=', row[2])]).id
                        # 查询员工
                        staff_id = self.env['hr.employee'].search([('name', '=', row[3])]).id
                        # 查询合同
                        contract = self.env['hr.employee'].search([('name', '=', row[3])]).contract_attributes

                        # 组装数据
                        combined_data = {
                            'employee_id': row[0],
                            'order_no': order_number_id,
                            'order_number': style_no_id,
                            'employee': staff_id,
                            'process_abbreviation': row[4],
                            'over_number': float(row[5]),
                            'standard_price': matched_record.standard_price,
                            'standard_time': matched_record.standard_time,
                            'group':row[8],
                            'date1': row[9],
                            'contract_type': contract
                        }
                        self.sudo().create(combined_data)

    @api.model
    def process_temporary_excel_data(self, data):
        """临时工excel导入"""

        decoded_data = base64.b64decode(data.split(',')[1])
        file_object = io.BytesIO(decoded_data)
        wb = openpyxl.load_workbook(file_object)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True, min_row=2):
            process_number = row[0]
            style_number = row[3]

            if process_number and style_number:
                temporary = self.env['temporary_workers_apply'].search([('process_no.employee_id', '=', process_number), ('style_number.style_number', '=', style_number)], limit=1)

                if not temporary:
                    raise ValidationError(
                        f"在工序单中没有搜索到Excel表中对应的工序号：{process_number}, 款号:{style_number} ！")
                else:
                    for temp in temporary:
                        # 查询订单号
                        order_number_id = self.env['sale_pro.sale_pro'].search([('order_number', '=', row[2])]).id
                        # 查询款号
                        style_no_id = self.env['ib.detail'].search([('style_number', '=', row[3])]).id
                        # 查询员工
                        staff_id = self.env['hr.employee'].search([('name', '=', row[4])]).id
                        # 查询合同
                        contract = self.env['hr.employee'].search([('name', '=', row[4])]).contract_attributes

                        # 组装数据
                        combined_data = {
                            'employee_id': row[0],
                            'order_no': order_number_id,
                            'order_number': style_no_id,
                            'employee': staff_id,
                            'process_abbreviation': row[5],
                            'over_number': row[6],
                            'standard_price': temp.apply_price,
                            'standard_time': temp.standard_time,
                            'group': row[7],
                            'date1': row[1],
                            'contract_type': contract
                        }

                        self.sudo().create(combined_data)




    def self_test(self):
        """检查工时工序件数与生产进度件数不符"""
        sum_by_key = defaultdict(int)
        record_by_key = {}  # 新增一个字典，保存订单号和款号对应的记录对象

        for record in self:
            key = (record.order_no.id, record.order_number.id)
            sum_by_key[key] += record.over_number
            record_by_key[key] = record  # 将订单号和款号对应的记录对象保存到 record_by_key 字典中

        for key, value in sum_by_key.items():
            order_number_id, style_number_id = key
            total_over_number = value

            record = record_by_key[key]

            ret_schedule = self.env['schedule_production'].search([
                ('order_number', '=', order_number_id),
                ('style_number', '=', style_number_id)
            ])

            total_qualified_stock = sum(production.qualified_stock for production in ret_schedule)

            if total_over_number > total_qualified_stock:
                record.non_conformance = True
            else:
                record.non_conformance = False

    def self_price(self):
        """检查工序价格"""
        for record in self:
            formal_worker = self.env['work.work'].search([('order_number', '=', record.order_number.id),
                                                          ('employee_id', '=', record.employee_id)])
            for formal in formal_worker:
                if formal.standard_price != record.standard_price:
                    record.non_price = True
                else:
                    record.non_price = False


    # 批量确认
    def approved_for_confirmation(self):
        for record in self:
            if record.state != "已确认":
                record.state = "已确认"
            else:
                raise ValidationError(f"已经是已确认状态了，无需再已确认！")

    # 批量未确认
    def batch_unconfirmed(self):
        for record in self:
            if record.state != "未确认":
                record.state = '未确认'
            else:
                raise ValidationError(f"已经是未确认状态了，无需再未确认！")

    # 设置临时工工资
    def set_casual_wage(self):

        casual_wage_obj = self.env["casual_wage"].sudo().search([
            ("dDate", "=", self.date1),
            ("employee_id", "=", self.employee.id)
        ])
        if casual_wage_obj:
            casual_wage_obj.sudo().set_data()
        else:
            new_obj = casual_wage_obj.create({
                "dDate": self.date1,
                "employee_id": self.employee.id,
                "contract_type": self.contract_type,
                "group": self.group
            })
            new_obj.sudo().set_data()


    # 设置个人效率
    def set_eff_eff(self):

        eff_eff_obj = self.env["eff.eff"].sudo().search([
            ("date", "=", self.date1),
            ("employee", "=", self.employee.id)
        ])
        if eff_eff_obj:
            eff_eff_obj.sudo().set_totle_eff()
        else:
            new_obj = eff_eff_obj.create({
                "date": self.date1,
                "employee": self.employee.id,
                "group": self.group
            })
            new_obj.sudo().set_totle_eff()


    # 设置组效率
    def set_group_efficiency(self):

        group_efficiency_obj = self.env["group.efficiency"].sudo().search([
            ("date", "=", self.date1),
            ("group", "=", self.group)
        ])
        if group_efficiency_obj:
            group_efficiency_obj.sudo().set_totle_eff()
        else:
            new_obj = group_efficiency_obj.create({
                "date": self.date1,
                "group": self.group
            })
            new_obj.sudo().set_totle_eff()


    # 设置计件工资
    def set_ji_jian(self):
        ji_jian_objs = self.env["ji.jian"].sudo().search([("date1", "=", self.date1), ("employee", "=", self.employee.id)])
        if ji_jian_objs:
            ji_jian_objs.sudo().set_cost()
        else:
            new_obj = ji_jian_objs.sudo().create({
                "date1": self.date1,
                "employee": self.employee.id,
                "contract_type": self.contract_type,
                "group": self.group
            })
            new_obj.sudo().set_cost()


    def set_piecerate_id(self):

        for record in self:

            cost_cost1_obj = self.env['cost.cost1'].sudo().create(
                {
                    'date1': record.date1,
                    'employee': record.employee.id,
                    'group': record.group,
                    'cost': record.standard_price * record.over_number
                }
            )
            record.piecerate_id = cost_cost1_obj.id



    @api.model
    def create(self, vals):

        # 创建员工信息
        memp_memp_obj = self.env['memp.memp'].sudo().create(
            {
                'date': vals['date1'],
                'employee': vals['employee'],
                'group': vals['group'],
            }
        )
        vals["made_employee_id"] = memp_memp_obj.id

        instance = super(on_work, self).create(vals)

        if instance.employee.is_it_a_temporary_worker == "临时工" or instance.employee.is_it_a_temporary_worker == "外包(计时)" or instance.employee.is_it_a_temporary_worker == "外包(计件)":
            # 设置临时工工资
            instance.set_casual_wage()

        # 设置个人效率表
        instance.set_eff_eff()

        # 设置组效率
        instance.set_group_efficiency()

        # 设置计件工资
        instance.set_ji_jian()

        # 设置工序工资
        instance.set_piecerate_id()

        return instance


    def unlink(self):

        for record in self:

            # 删除工序工资
            record.piecerate_id.sudo().unlink()
            # 删除员工表
            record.made_employee_id.sudo().unlink()

            dDate = record.date1    # 日期
            employee_id = record.employee.id    # 员工id
            type_of_work = record.employee.is_it_a_temporary_worker     # 工种
            gGroup = record.group   # 组别

            super(on_work, record).unlink()

            if type_of_work == "临时工" or type_of_work == "外包(计时)" or type_of_work == "外包(计件)":

                # 删除时，同步临时工，外包工资
                casual_wage_obs = self.env["casual_wage"].sudo().search([
                    ("dDate", "=", dDate),
                    ("employee_id", "=", employee_id)
                ])
                if casual_wage_obs:
                    casual_wage_obs.sudo().set_data()


            # 删除时，同步个人效率表
            eff_eff_obj = self.env["eff.eff"].sudo().search([
                ("date", "=", dDate),
                ("employee", "=", employee_id)
            ])
            if eff_eff_obj:
                eff_eff_obj.sudo().set_totle_eff()


            # 删除时，同步组效率
            group_efficiency_obj = self.env["group.efficiency"].sudo().search([
                ("date", "=", dDate),
                ("group", "=", gGroup)
            ])
            if group_efficiency_obj:
                group_efficiency_obj.sudo().set_totle_eff()


            # 删除时，同步计件工资
            ji_jian_objs = self.env["ji.jian"].sudo().search([
                ("date1", "=", dDate),
                ("employee", "=", employee_id),
            ])
            if ji_jian_objs:
                ji_jian_objs.sudo().set_cost()


        return super(on_work, self).unlink()


